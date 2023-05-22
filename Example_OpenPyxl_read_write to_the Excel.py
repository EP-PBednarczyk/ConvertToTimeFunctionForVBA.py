# ToDo example from the www youtube -> how to use:
#  https://www.youtube.com/watch?v=gY5oQIBjIg8 'How to change Date 
#  Formats in Excel using Python | From any Format to any other
#  'czas_pracy_1.4.28_Pawel_Bednarczyk.xlsm' file excel

import openpyxl as _opPyxl

file = _opPyxl.load_workbook('Exercise_read_write_OpenPyxl.xlsx')    # 'czas_pracy_1.4.28_Pawel_Bednarczyk.xlsm')

sheets = file.sheetnames
# print(sheet)  result: ['May_2023']
sheet = file[sheets[0]]
rows = sheet.max_row
# print(rows)   result: 21 rows

for i in range(3, 34):
# for i in range(2, rows):
    cell = 'A' + f'{i}'
    # print(cell)
    date = sheet[cell].value
    print(date)

    # results:
    # 2023-05-01 00:00:00
    # 2023-05-02 00:00:00
    # 2023-05-03 00:00:00
    # 2023-05-04 00:00:00
    # 2023-05-05 00:00:00
    # 2023-05-06 00:00:00
    # 2023-05-07 00:00:00
    # 2023-05-08 00:00:00
    # 2023-05-09 00:00:00
    # 2023-05-10 00:00:00
    # 2023-05-11 00:00:00
    # 2023-05-12 00:00:00
    # 2023-05-13 00:00:00
    # 2023-05-14 00:00:00
    # 2023-05-15 00:00:00
    # 2023-05-16 00:00:00
    # 2023-05-17 00:00:00
    # 2023-05-18 00:00:00
    # 2023-05-19 00:00:00
    # 2023-05-20 00:00:00
    # 2023-05-21 00:00:00
    # 2023-05-22 00:00:00
    # 2023-05-23 00:00:00
    # 2023-05-24 00:00:00
    # 2023-05-25 00:00:00
    # 2023-05-26 00:00:00
    # 2023-05-27 00:00:00
    # 2023-05-28 00:00:00
    # 2023-05-29 00:00:00
    # 2023-05-30 00:00:00
    # 2023-05-31 00:00:00

    #list_data = date.split("/")
    #print(list_data)
    ## results:
    ## error  File "<input>", line 6, in <module>
    ## AttributeError: 'datetime.datetime' object has no attribute 'split'