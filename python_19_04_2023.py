import string
import openpyxl
import pandas as pd  # to the excel
# windowmessage
import ctypes  # An included library with Python install.
import warnings

# openpyxl or XlsxWriter to write to .xlsx files
# https://openpyxl.readthedocs.io/en/stable/
import openpyxl


# ------------------------------------------------------------------------------------- #
# import warnings
# from openpyxl import *  # library to the excel - GNU license
# from openpyxl import Workbook, load_workbook  # library to the excel - GNU license


# to the test
# _input_included_time_wrong_symbol = ["abc#@!?efg;:*$**?***08:00:00"]  # test

# wb = load_workbook('czas_pracy_1.3.24_Pawel_Bednarczyk_IV_2023_04-04.xlsm', keep_vba=True)
# ws = wb.active
# warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
# file_name = 'czas_pracy_1.4.12_Pawel_Bednarczyk_IV_2023_04-14.xlsm'


#@ToDo - Excel musi byc zamkniety???
try:
    # load xls file ->from pandass
    # nie dziala: d = pd.read_csv(r'czas_pracy_1.4.12_Pawel_Bednarczyk_IV_2023_04-14.csv', header=None)
    d = pd.read_excel(r'czas_pracy_1.4.12_Pawel_Bednarczyk_IV_2023_04-14.xlsm', index_col=0)

    # from openpyxl
    # d = openpyxl.load_workbook(file_name, data_only = True ,read_only = False, keep_vba = True)
    warnings.simplefilter(action='ignore')
except FileNotFoundError:
    print("File could not be found.")

# ???????????????????????????????????????????
## @ToDo - sprawdzic przyklad .... ?
## Read Excel and select a single cell (and make it a header for a column)
# filename = "czas_pracy_1.4.12_Pawel_Bednarczyk_IV_2023_04-14.xlsm"
# data = pd.read_excel(filename, 'Sheet2', index_col=None, usecols = "C", header = 10, nrows=0)
## Will return a "list" of 1 header(s) and no data. Then isolate that header:

## Extract a value from a list (list of headers)
# data = data.columns.values[0]
# print (data)
# ???????????????????????????????????????????


_input_included_time_wrong_symbol = d.at[12, 12]  # .at[numer wiersza, numer kolumny z excela]
_input_included_time_wrong_symbol.append(d.at[12, 13])
_input_included_time_wrong_symbol.append(d.at[12, 14])
_input_included_time_wrong_symbol.append(d.at[12, 15])

print(_input_included_time_wrong_symbol)


def convert_to_time(_input_included_time_wrong_symbol):
    # convert string to Time 00:00:00 [hours:minutes:seconds]'
    # remove chars(words) from times
    # string_with_time.remove()  # remove chars(words) from times
    # string_with_time = string_with_time.translate({ord(c): None for c in '***'})

    # ASCII: 33 to 47 - !,.../
    str_lower_word = list(string.ascii_lowercase)
    str_upper_word = list(string.ascii_uppercase)

    # special owns:
    _prohibited_characters = ['*', '**', '***', '****', '!', '@', '#',
                              '$', '%', '^', '&', '(', ')', '-', '+',
                              '_', '=', '[', ']', '{' '}', ';', "'",
                              '"', '\\', '|', '<', '>', ',', '.',
                              '/', '?', '|', '`', '~'] + str_lower_word + str_upper_word
    #  unecessary marks going to remove
    _fixed_string: str = ''.join(c for c in _input_included_time_wrong_symbol[0] if c not in _prohibited_characters)

    # wsadzenie z powrotem do excela
    d.at[12, 12] = _fixed_string[0]
    d.at[12, 13] = _fixed_string[1]
    d.at[12, 14] = _fixed_string[2]
    d.at[12, 15] = _fixed_string[3]
    # -----------------to the debug ---------------------------------------
    # windowd message
    ctypes.windll.user32.MessageBoxW(0, "Your text  " + d, "Your title", 1)

    # -------------------------------------------------------------------

    return

# ToDo issue when case ":" ":08:00:00    08:00:00:, 08:0:0:0:0 etc"
# if c == ":"

# if __name__ == '__main__':
#    print(convert_to_time())
# warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
