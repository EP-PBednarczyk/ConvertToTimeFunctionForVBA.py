import string
import sys
import pandas as pd  # to the excel
import ctypes  # An included library with Python install.
import warnings
import openpyxl    # example-> how to use: https://www.youtube.com/watch?v=gY5oQIBjIg8 'How to change Date Formats in Excel using Python | From any Format to any other'


# openpyxl or XlsxWriter to write to .xlsx files, library to the excel - GNU license
# https://openpyxl.readthedocs.io/en/stable/
# from openpyxl import *  # library to the excel - GNU license
# from openpyxl import Workbook, load_workbook  # library to the excel - GNU license

# to the test
# _input_included_time_wrong_symbol = ["abc#@!?efg;:*$**?***08:00:00"]  # test
# wb = load_workbook('czas_pracy_1.4.28_Pawel_Bednarczyk_IV_2023_04-28.xlsm', keep_vba=True)
# ws = wb.active
# warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
# file_name = 'czas_pracy_1.4.12_Pawel_Bednarczyk_IV_2023_04-14.xlsm'

try:
    # load xls file ->from pandas
    data = pd.read_excel(r'czas_pracy_1.4.28_Pawel_Bednarczyk.xlsm', index_col=0, sheet_name=None)
    print(data)
    # from openpyxl
    # d = openpyxl.load_workbook(file_name, data_only = True ,read_only = False, keep_vba = True)
    warnings.simplefilter(action='ignore')  # ignore warnings
except FileNotFoundError:
    print("File could not be found.")
# ----------------------
# data
# result: [19 rows x 38 columns]
# data.get("Unnamed: 1")[0]
# result: 'Czas pracy 1.4.12'

# data matrix from excel (read celles where exist date):
# [19 rows x 38 columns] in excel: from 2 to 20 row ; columns B:AL

# >>> data.keys(),
# result: dict_keys(['April_2023'])
# ----------------------------------
# data [2]
# results: matrix ... [19 rows x 37 columns]}
# data.get('April_2023').values
# results:
# array([['Czas pracy 1.4.12', nan, nan, nan, nan, nan, nan, nan, nan, nan,
#         nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan,
#         nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan,
#         nan],
#        ['ROK', 2023, 'MMXXIII', nan, nan, nan, 4, nan, nan, nan,
#         'Czas modyfikacji:', nan, nan, nan, nan,
#         datetime.datetime(2023, 4, 26, 13, 21, 32, 330000), nan, nan,
#         nan, nan, 'ścieżka katalogu:', nan, nan,
#         'C:\\Users\\pbednarczyk\\Documents\\', nan, nan, nan, nan, nan,
#         'wersja excela:', nan, '15.0', nan, 'środowisko\n operacyjne:',
#         nan, nan, 'pcdos'],
#        ['Miesiąc', 4, 'Kwiecień', nan, nan, nan, nan, nan, nan, nan, nan,
#         nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan,
#         nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan],
#        ['Nr. Tygodnia', nan, 'Tydzień 14 roku', nan, nan, nan, nan, nan,
#         nan, 'Tydzień 15 roku', nan, nan, nan, nan, nan, nan,
#         'Tydzień 16 roku', nan, nan, nan, nan, nan, nan,
#         'Tydzień 17 roku', nan, nan, nan, nan, nan, nan,
#         'Tydzień 18 roku', nan, nan, nan, nan, nan, nan],
#        ['                                \n                                   Dni\n\nNazwa projektu',
#         ' nr zlecenia\n \n        godz. pracy w tyg.',
#         datetime.datetime(2023, 4, 3, 0, 0),
#         datetime.datetime(2023, 4, 4, 0, 0),
#         datetime.datetime(2023, 4, 5, 0, 0),
#         datetime.datetime(2023, 4, 6, 0, 0),
#         datetime.datetime(2023, 4, 7, 0, 0), 'suma tyg.',
#         'suma tyg.+ poprzednie\ntyg. miesiąca',
#         datetime.datetime(2023, 4, 10, 0, 0),
#         datetime.datetime(2023, 4, 11, 0, 0),
#         datetime.datetime(2023, 4, 12, 0, 0),
#         datetime.datetime(2023, 4, 13, 0, 0),
#         datetime.datetime(2023, 4, 14, 0, 0), 'suma tyg.',
#         'suma tyg.+ poprzednie\ntyg. miesiąca',
#         datetime.datetime(2023, 4, 17, 0, 0),
#         datetime.datetime(2023, 4, 18, 0, 0),
#         datetime.datetime(2023, 4, 19, 0, 0),
#         datetime.datetime(2023, 4, 20, 0, 0),
#         datetime.datetime(2023, 4, 21, 0, 0), 'suma tyg.',
#         'suma tyg.+ poprzednie\ntyg. miesiąca',
#         datetime.datetime(2023, 4, 24, 0, 0),
#         datetime.datetime(2023, 4, 25, 0, 0),
#         datetime.datetime(2023, 4, 26, 0, 0),
#         datetime.datetime(2023, 4, 27, 0, 0),
#         datetime.datetime(2023, 4, 28, 0, 0), 'suma tyg.',
#         'suma tyg.+ poprzednie\ntyg. miesiąca',
#         datetime.datetime(2023, 5, 1, 0, 0),
#         datetime.datetime(2023, 5, 2, 0, 0),
#         datetime.datetime(2023, 5, 3, 0, 0),
#         datetime.datetime(2023, 5, 4, 0, 0),
#         datetime.datetime(2023, 5, 5, 0, 0), 'suma tyg.',
#         'suma tyg.+ poprzednie\ntyg. miesiąca'],
#        ['∑ godz. w danym dniu tyg.', nan, datetime.time(8, 0),
#         datetime.time(8, 0), datetime.time(8, 0), datetime.time(8, 0),
#         datetime.time(8, 0), 'godz.:min', 'godz:min:sek',
#         datetime.time(0, 0), datetime.time(8, 0), datetime.time(8, 0),
#         datetime.time(8, 0), datetime.time(8, 0), 'godz.:min',
#         'godz:min:sek', datetime.time(8, 0), datetime.time(8, 0),
#         datetime.time(8, 0), datetime.time(8, 0), datetime.time(8, 0),
#         'godz.:min', 'godz:min:sek', datetime.time(8, 0),
#         datetime.time(8, 0), datetime.time(1, 0), datetime.time(0, 0),
#         datetime.time(0, 0), 'godz.:min', 'godz:min:sek',
#         datetime.time(0, 0), datetime.time(0, 0), datetime.time(0, 0),
#         datetime.time(0, 0), datetime.time(0, 0), 'godz.:min',
#         'godz:min:sek'],
#        ['TAURON Ciepło - Testy SGU dla Bielsko-Biała zakład EC1 i EC2',
#         '200.1.22.2314.3277.4', nan, nan, nan, nan, nan,
#         datetime.time(0, 0), datetime.time(0, 0), 'święto', nan, nan,
#         nan, nan, datetime.time(0, 0), datetime.time(0, 0), nan, nan,
#         nan, nan, nan, datetime.time(0, 0), datetime.time(0, 0), nan,
#         nan, nan, nan, nan, datetime.time(0, 0), datetime.time(0, 0),
#         nan, nan, nan, nan, nan, datetime.time(0, 0),
#         datetime.time(0, 0)],
#        ['Pge El. Turów - Testy Wewn. I Odb. Rp I Rw Na Bl.1-5 I 7\n',
#         '200.1.22.9155.2423.4', nan, nan, nan, nan, nan,
#         datetime.time(0, 0), datetime.time(0, 0), 'święto', nan, nan,
#         nan, nan, datetime.time(0, 0), datetime.time(0, 0), nan, nan,
#         nan, nan, nan, datetime.time(0, 0), datetime.time(0, 0), nan,
#         nan, nan, nan, nan, datetime.time(0, 0), datetime.time(0, 0),
#         nan, nan, nan, nan, nan, datetime.time(0, 0),
#         datetime.time(0, 0)],
#        ['Testy Samostartu W Ew Niedzica', '200.1.23.1062.3365.4', nan,
#         nan, nan, nan, nan, datetime.time(0, 0), datetime.time(0, 0),
#         'święto', nan, nan, nan, nan, datetime.time(0, 0),
#         datetime.time(0, 0), nan, nan, nan, nan, nan,
#         datetime.time(0, 0), datetime.time(0, 0), nan, nan, nan, nan,
#         nan, datetime.time(0, 0), datetime.time(0, 0), nan, nan, nan,
#         nan, nan, datetime.time(0, 0), datetime.time(0, 0)],
#        ['Wydziałowe - inne prace zlecone przez kierownika', nan,
#         datetime.time(8, 0), datetime.time(7, 0), datetime.time(8, 0),
#         nan, nan, datetime.time(23, 0), datetime.time(23, 0), 'święto',
#         datetime.time(8, 0), datetime.time(8, 0), datetime.time(8, 0),
#         datetime.time(8, 0), datetime.datetime(1900, 1, 1, 8, 0),
#         datetime.datetime(1900, 1, 2, 7, 0), datetime.time(8, 0),
#         datetime.time(8, 0), datetime.time(8, 0), datetime.time(8, 0),
#         datetime.time(8, 0), datetime.datetime(1900, 1, 1, 16, 0),
#         datetime.datetime(1900, 1, 3, 23, 0), datetime.time(7, 0),
#         datetime.time(5, 0), nan, nan, nan, datetime.time(12, 0),
#         datetime.datetime(1900, 1, 4, 11, 0), nan, nan, nan, nan, nan,
#         datetime.time(0, 0), datetime.datetime(1900, 1, 4, 11, 0)],
#        ['*praca zdalna (prace zlecone przez kierownika),\n**dojazd do zakładu pracy',
#         nan, nan, nan, nan, nan, nan, datetime.time(0, 0),
#         datetime.time(0, 0), 'święto', nan, nan, nan, '**00:00:00', nan,
#         nan, nan, nan, nan, nan, nan,
#         "=SUMA(convert_to_time(WK[-5]):convert_to_time(WK[-1])))'", nan,
#         nan, nan, nan, nan, nan,
#         '"=SUMA(convert_to_time(WK[-5]:WK[-1]))"', nan, nan, nan, nan,
#         nan, nan, '"=SUMA(convert_to_time(WK[-5]:WK[-1]))"', nan],
#        ['Urlop Wypoczynkowy', nan, nan, nan, nan, datetime.time(8, 0),
#         datetime.time(8, 0), datetime.time(16, 0), datetime.time(16, 0),
#         'święto', nan, nan, nan, nan, datetime.time(0, 0), nan, nan, nan,
#         nan, nan, nan, datetime.time(0, 0), datetime.time(0, 0), nan,
#         nan, nan, nan, nan, datetime.time(0, 0), datetime.time(0, 0),
#         nan, nan, nan, nan, nan, datetime.time(0, 0),
#         datetime.time(0, 0)],
#        ['Szkolenia / Konferencje', nan, nan, datetime.time(1, 0), nan,
#         nan, nan, datetime.time(1, 0), datetime.time(1, 0), 'święto',
#         nan, nan, nan, nan, datetime.time(0, 0), datetime.time(1, 0),
#         nan, nan, nan, nan, nan, datetime.time(0, 0),
#         datetime.time(1, 0), datetime.time(1, 0), datetime.time(3, 0),
#         datetime.time(1, 0), nan, nan, datetime.time(5, 0),
#         datetime.time(6, 0), nan, nan, nan, nan, nan,
#         datetime.time(0, 0), datetime.time(6, 0)],
#        [nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan,
#         nan, nan, nan, 'legenda czasu: 0.25 = 15 minut', nan, nan, nan,
#         nan, nan, nan, ' 0.17 = 10 minut', nan, nan, nan, nan, nan, nan,
#         nan, nan, nan, nan, nan, nan, nan],
#        ['∑ godz. w tygodniu', nan, nan, nan, nan, nan, nan,
#         datetime.datetime(1900, 1, 1, 16, 0), nan, nan, nan, nan, nan,
#         nan, nan, nan, nan, nan, nan, nan, nan,
#         datetime.datetime(1900, 1, 1, 16, 0), nan, nan, nan, nan, nan,
#         nan, datetime.time(17, 0), nan, nan, nan, nan, nan, nan,
#         datetime.time(0, 0), nan],
#        ['∑ godz. bieżącego tyg.\n i poprzednich tyg.', nan, nan, nan,
#         nan, nan, nan, nan, datetime.datetime(1900, 1, 1, 16, 0), nan,
#         nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan,
#         nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan,
#         nan],
#        [nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan,
#         nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan,
#         nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan],
#        ['*VBA script was written based on framework from book:
#        "Excel z elementami VBA w firmie" - Sergiusz Flanczewski.',
#         nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan,
#         nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan,
#         nan, nan, nan, nan, nan, nan, nan, nan, nan, nan],
#        ['mailto:pbednarczyk@energopomiar.com', nan, nan, nan, nan, nan,
#         nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan,
#         nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan,
#         nan, nan, nan, nan, nan]], type=object)
# kk.get('April_2023').values[0][0]   -> (Excel B2)
# 'Czas pracy 1.4.12'
# kk.get('April_2023').values[9][2]  -> (Excel D11)
# results: datetime.time(8, 0)  -> w Excelu 08:00:00
# _input_included_time_wrong_symbol   ->  '**00:00:00'

# -----------------window text for the print and test-------
_input_included_time_wrong_symbol = data.get('May_2023').values[10][13]  # (excel cell O12 'O' as Ola)
print(_input_included_time_wrong_symbol)

ctypes.windll.user32.MessageBoxW(0, "Your text ", "Your title", 1)


# --------------------------------------------------------

def convert_to_time(_input_included_time_wrong_symbol):

    # convert string to Time 00:00:00 [hours:minutes:seconds]'
    # remove chars(words) from times
    # string_with_time.remove() -> remove chars(words) from times
    # string_with_time = string_with_time.translate({ord(c): None for c in '***'})
    # ASCII: 33 to 47 - !,.../

    str_lower_word = list(string.ascii_lowercase)
    str_upper_word = list(string.ascii_uppercase)

    # special owns:
    _prohibited_characters = ['*', '**', '***', '****', '!', '@', '#',
                              '$', '%', '^', '&', '(', ')', '-', '+',
                              '_', '=', '[', ']', '{' '}', ';', "'",
                              '"', '\\', '|', '<', '>', ',', '.',
                              '/', '?', '|', '`', '~'] + str_lower_word + str_upper_word
    #  unnecessary marks going to remove
    _fixed_string: str = ''.join(c for c in _input_included_time_wrong_symbol if c not in _prohibited_characters)

    # wsadzenie z powrotem do excela

    # -----------------to the debug ---------------------------------------
    # window message with button
    ctypes.windll.user32.MessageBoxW(0, "Your text from cell " + _input_included_time_wrong_symbol +
                                     "... after remove unnecessary signs: " + _fixed_string
                                     , "Your title", 1)
    # -------------------------------------------------------------------

    # cell print in python. From pandas object   <- this cells I wpuld like to update
    ###################################################
    # example -> write update Excel
    # https: // stackoverflow.com / questions / 47891444 / how - can - i - update - my - dataframe - in -pandas - and -export - out - to - excel
    ##################################################################
    #   import pandas as pd
    # df = pd.read_excel(my_file, sheet_name='Sheet1')
    #
    #   dept = df['department']
    #   resource = df['resource']
    #   start_appointment = df['start appointment']
    #
    #   def diagnostic():  # Check Diagnostic Breast scheduled appointments
    #       for i in range(10):
    #           minutes = str(start_appointment[i])[14:16]
    #           hour = str(start_appointment[i])[11:13]
    #           if minutes == '15' and (
    #                   hour == '8' or hour == '9' or hour == '10' or hour == '11'
    #             or hour == '13' or hour == '14' or hour == '15') and (
    #             resource[i] == 'BIDAG1' or resource[i] == 'BDIAG2' or
    #             resource[i] == 'BDIAG3'):
    #         df.update['resource'][i] = 'ZBMDX3'
    #     elif minutes == '00' and (hour == '8' or hour == '9' or hour == '10' or
    #             hour == '11' or hour == '13' or hour == '14' or hour == '15')
    #             and (resource[i] == 'BIDAG1' or resource[i] == 'BDIAG2' or
    #             resource[i] == 'BDIAG2'):
    #         df.update['resource'][i] = 'ZBMDX2'
    #     elif minutes == '45' and (
    #             hour == '7' or hour == '8' or hour == '9' or hour == '10' or
    #             hour == '12' or hour == '13' or hour == '14') and (
    #             resource[i] == 'BIDAG1' or resource[i] == 'BDIAG2' or
    #             resource[i] == 'BDIAG1'):
    #         df.update['resource'][i] = 'ZBMDX1'
    #     elif minutes == '30' and (hour == '8' or hour == '9' or hour == '10' or
    #             hour == '13' or hour == '14') and (
    #             resource[i] == 'BIDAG1' or resource[i] == 'BDIAG2' or
    #             resource[i] == 'BDIAG1'):
    #         df.update['resource'][i] = 'ZBMDX4'
    #   diagnostic()
    #
    # # Specify a writer
    # writer = pd.ExcelWriter('C:\\Users\user_name\Desktop\Python 3\Python_Output.xlsx', engine='xlsxwriter')
    #
    # # Write your DataFrame to a file
    # df.to_excel(writer, 'Sheet1')
    #
    # # Save the result
    # writer.save()
    #---------------------------------------------------------------------------------------


    print(list(data['May_2023']['Unnamed: 14'])[10])
    wrong_data = list(data['May_2023']['Unnamed: 14'])[
        10]  # print  komórka O12 to N(14) liczone od kolumny B(1), wiersze liczone od 3 początek tabeli a indeksowanie w pythonie zaczyna się od 0) *00:00:00
    # '**00:00:00'
    list(data.update['May_2023']['Unnamed: 14'])[10] = _fixed_string

    return




input("hit any key to close")
sys.exit()
# ------------------------????????--ToDo---------------------
# ToDo issue when case ":" ":08:00:00    08:00:00:, 08:0:0:0:0 etc"
# if c == ":"
# if __name__ == '__main__':
#    print(convert_to_time())
# warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
